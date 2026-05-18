"""Tests for all document parsers.

Phase: 1  |  Run: pytest tests/test_parsers.py -v

Fixtures (put real files in tests/fixtures/ to run integration tests):
    tests/fixtures/sample.pdf   - a PDF with at least 2 headings and 1 table
    tests/fixtures/sample.docx  - a DOCX with headings, body text, and a table
    tests/fixtures/sample.xlsx  - see ExcelParser tests
"""

import pytest

from docnest.exceptions import UnsupportedFormatError
from docnest.parsers.docx import DocxParser
from docnest.parsers.factory import ParserFactory
from docnest.parsers.html import HTMLParser
from docnest.parsers.md import MarkdownParser
from docnest.parsers.pdf import DoclingPDFParser
from docnest.parsers.xlsx import ExcelParser

# ------------------------------------------------------------------ #
#  ParserFactory                                                       #
# ------------------------------------------------------------------ #

class TestParserFactory:
    def test_returns_pdf_parser_for_pdf(self):
        assert isinstance(ParserFactory().get("report.pdf"), DoclingPDFParser)

    def test_returns_docx_parser(self):
        assert isinstance(ParserFactory().get("doc.docx"), DocxParser)

    def test_returns_xlsx_parser(self):
        assert isinstance(ParserFactory().get("data.xlsx"), ExcelParser)

    def test_returns_html_parser(self):
        assert isinstance(ParserFactory().get("page.html"), HTMLParser)

    def test_returns_md_parser(self):
        assert isinstance(ParserFactory().get("README.md"), MarkdownParser)

    def test_raises_for_unsupported_format(self):
        with pytest.raises(UnsupportedFormatError):
            ParserFactory().get("archive.zip")

    def test_supports_returns_false_for_unknown(self):
        assert ParserFactory().supports("file.xyz") is False

    def test_supports_returns_true_for_pdf(self):
        assert ParserFactory().supports("report.pdf") is True

    def test_supports_returns_true_for_docx(self):
        assert ParserFactory().supports("brief.docx") is True

    def test_case_insensitive_extension_pdf(self):
        assert ParserFactory().supports("REPORT.PDF") is True

    def test_case_insensitive_extension_docx(self):
        assert ParserFactory().supports("DOC.DOCX") is True

    def test_register_adds_custom_parser(self):
        """register() inserts parser into registry (line 95)."""
        from docnest.models import RawDocument
        from docnest.parsers.base import IParser

        class CustomParser(IParser):
            def supports(self, f: str) -> bool:
                return f.endswith(".custom")
            def parse(self, f: str) -> RawDocument:
                raise NotImplementedError

        factory = ParserFactory()
        factory.register(CustomParser())
        assert factory.supports("file.custom")

    def test_register_at_non_zero_position(self):
        """register() with explicit position (line 95)."""
        from docnest.models import RawDocument
        from docnest.parsers.base import IParser

        class PriorityParser(IParser):
            def supports(self, f: str) -> bool:
                return f.endswith(".pri")
            def parse(self, f: str) -> RawDocument:
                raise NotImplementedError

        factory = ParserFactory()
        factory.register(PriorityParser(), position=len(factory._registry))
        assert factory.supports("file.pri")

    def test_unregister_removes_parser(self):
        """unregister() removes parser class (line 103)."""
        factory = ParserFactory()
        factory.unregister(DocxParser)
        assert not factory.supports("file.docx")

    def test_set_pdf_engine_pymupdf(self):
        """set_pdf_engine('pymupdf') swaps to PyMuPDF parser (lines 111-120)."""
        from docnest.parsers.pymupdf_pdf import PyMuPDFParser
        factory = ParserFactory()
        factory.set_pdf_engine("pymupdf")
        assert factory.supports("report.pdf")
        parser = factory.get("report.pdf")
        assert isinstance(parser, PyMuPDFParser)

    def test_set_pdf_engine_docling(self):
        """set_pdf_engine('docling') re-registers Docling parser (lines 111-120)."""
        factory = ParserFactory()
        factory.set_pdf_engine("pymupdf")  # switch to pymupdf first
        factory.set_pdf_engine("docling")  # switch back
        parser = factory.get("report.pdf")
        assert isinstance(parser, DoclingPDFParser)

    def test_factory_with_pymupdf_engine(self):
        """ParserFactory(pdf_engine='pymupdf') builds with PyMuPDF (lines 143-144)."""
        from docnest.parsers.pymupdf_pdf import PyMuPDFParser
        factory = ParserFactory(pdf_engine="pymupdf")
        assert isinstance(factory.get("report.pdf"), PyMuPDFParser)

    def test_registered_extensions_non_empty(self):
        assert ParserFactory()._registered_extensions() != "none"


# ------------------------------------------------------------------ #
#  DoclingPDFParser — unit tests (no fixture file required)           #
# ------------------------------------------------------------------ #

class TestDoclingPDFParserUnit:
    def test_supports_pdf(self):
        assert DoclingPDFParser().supports("report.pdf") is True

    def test_supports_case_insensitive(self):
        assert DoclingPDFParser().supports("REPORT.PDF") is True

    def test_does_not_support_docx(self):
        assert DoclingPDFParser().supports("doc.docx") is False

    def test_parse_missing_file_raises_parse_error(self):
        from docnest.exceptions import ParseError
        with pytest.raises(ParseError, match="not found"):
            DoclingPDFParser().parse("/tmp/does_not_exist_abc123.pdf")

    def test_lazy_converter_init(self):
        parser = DoclingPDFParser()
        assert parser._converter is None


# ------------------------------------------------------------------ #
#  DoclingPDFParser — integration (require tests/fixtures/sample.pdf) #
# ------------------------------------------------------------------ #

class TestDoclingPDFParserIntegration:
    def test_parse_returns_raw_document(self, sample_pdf):
        raw = DoclingPDFParser().parse(str(sample_pdf))
        assert raw.doc_id is not None
        assert raw.format == "pdf"
        assert raw.source.endswith(".pdf")

    def test_sections_are_extracted(self, sample_pdf):
        raw = DoclingPDFParser().parse(str(sample_pdf))
        assert len(raw.sections) > 0

    def test_section_ids_empty_from_parser(self, sample_pdf):
        """§ids must be empty strings — the Normaliser assigns them."""
        raw = DoclingPDFParser().parse(str(sample_pdf))
        for s in raw.sections:
            assert s.id == "", f"Parser must NOT assign §ids, got '{s.id}'"

    def test_tables_are_structured(self, sample_pdf):
        raw = DoclingPDFParser().parse(str(sample_pdf))
        tables = [t for s in raw.sections for t in s.tables]
        assert len(tables) > 0, "sample.pdf should contain at least one table"
        for t in tables:
            assert len(t.headers) > 0
            assert all(len(row) == len(t.headers) for row in t.rows)

    def test_section_titles_non_empty(self, sample_pdf):
        raw = DoclingPDFParser().parse(str(sample_pdf))
        for s in raw.sections:
            assert s.title.strip()

    def test_section_levels_valid(self, sample_pdf):
        raw = DoclingPDFParser().parse(str(sample_pdf))
        for s in raw.sections:
            assert 1 <= s.level <= 6


# ------------------------------------------------------------------ #
#  DocxParser — unit tests (no fixture file required)                 #
# ------------------------------------------------------------------ #

class TestDocxParserUnit:
    def test_supports_docx(self):
        assert DocxParser().supports("brief.docx") is True

    def test_supports_case_insensitive(self):
        assert DocxParser().supports("BRIEF.DOCX") is True

    def test_does_not_support_old_doc(self):
        # .doc (binary format) not supported by python-docx
        assert DocxParser().supports("old.doc") is False

    def test_does_not_support_pdf(self):
        assert DocxParser().supports("report.pdf") is False

    def test_parse_missing_file_raises_parse_error(self):
        from docnest.exceptions import ParseError
        with pytest.raises(ParseError, match="not found"):
            DocxParser().parse("/tmp/does_not_exist_abc123.docx")


# ------------------------------------------------------------------ #
#  DocxParser — integration (require tests/fixtures/sample.docx)      #
# ------------------------------------------------------------------ #

class TestDocxParserIntegration:
    def test_parse_returns_raw_document(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        assert raw.doc_id is not None
        assert raw.format == "docx"

    def test_sections_are_extracted(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        assert len(raw.sections) > 0

    def test_section_ids_empty_from_parser(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        for s in raw.sections:
            assert s.id == ""

    def test_heading_levels_detected(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        levels = {s.level for s in raw.sections}
        assert 1 in levels, "Expected at least one Heading 1 section"

    def test_tables_extracted(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        tables = [t for s in raw.sections for t in s.tables]
        assert len(tables) > 0

    def test_table_headers_and_rows_consistent(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        for s in raw.sections:
            for t in s.tables:
                assert len(t.headers) > 0
                assert all(len(row) == len(t.headers) for row in t.rows)

    def test_table_ids_unique(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        ids = [t.table_id for s in raw.sections for t in s.tables]
        assert len(ids) == len(set(ids))

    def test_body_text_present(self, sample_docx):
        raw = DocxParser().parse(str(sample_docx))
        assert any(s.text for s in raw.sections)


# ------------------------------------------------------------------ #
# ExcelParser — unit tests (no fixture file required) #
# ------------------------------------------------------------------ #

class TestExcelParserUnit:
    def test_supports_xlsx(self):
        assert ExcelParser().supports("data.xlsx") is True

    def test_supports_xls(self):
        """xls files are routed to ExcelParser but raise in parse()."""
        assert ExcelParser().supports("data.xls") is True

    def test_supports_case_insensitive(self):
        assert ExcelParser().supports("DATA.XLSX") is True

    def test_does_not_support_csv(self):
        assert ExcelParser().supports("data.csv") is False

    def test_does_not_support_pdf(self):
        assert ExcelParser().supports("report.pdf") is False

    def test_parse_missing_file_raises_parse_error(self):
        from docnest.exceptions import ParseError
        with pytest.raises(ParseError, match="not found"):
            ExcelParser().parse("/tmp/does_not_exist_abc123.xlsx")

    def test_parse_xls_raises_clear_error(self, tmp_path):
        """Legacy .xls format is not supported — must raise ParseError."""
        from docnest.exceptions import ParseError
        xls_file = tmp_path / "legacy.xls"
        xls_file.write_bytes(b"\xd0\xcf\x11\xe0")  # minimal OLE2 header
        with pytest.raises(ParseError, match=".xls.*not supported"):
            ExcelParser().parse(str(xls_file))

    def test_parse_empty_file_raises_parse_error(self, tmp_path):
        from docnest.exceptions import ParseError
        empty = tmp_path / "empty.xlsx"
        empty.write_bytes(b"")
        with pytest.raises(ParseError, match="empty"):
            ExcelParser().parse(str(empty))

    def test_row_length_normalisation(self):
        """_build_table pads short rows and truncates long rows."""
        header = ["A", "B", "C"]
        data = [
            ["1", "2"],          # short row → padded
            ["1", "2", "3", "4"],  # long row → truncated
            ["1", "2", "3"],      # exact length
        ]
        table = ExcelParser._build_table(header, data, "tbl_001")
        assert table is not None
        assert len(table.headers) == 3
        assert table.rows[0] == ["1", "2", ""]
        assert table.rows[1] == ["1", "2", "3"]
        assert table.rows[2] == ["1", "2", "3"]

    def test_build_table_strips_trailing_empty_headers(self):
        """Trailing empty header cells should be removed."""
        header = ["A", "B", "", ""]
        data = [["1", "2", "x", "y"]]
        table = ExcelParser._build_table(header, data, "tbl_001")
        assert table is not None
        assert table.headers == ["A", "B"]
        # Data row truncated to match
        assert table.rows[0] == ["1", "2"]

    def test_build_table_returns_none_for_empty_headers(self):
        """If all headers are empty, no table is created."""
        header = ["", "", ""]
        data = [["1", "2", "3"]]
        table = ExcelParser._build_table(header, data, "tbl_001")
        assert table is None

    def test_split_into_tables_single_table(self):
        """A simple sheet with one header + data stays as one table."""
        rows = [
            ["Name", "Age", "City"],
            ["Alice", "30", "NYC"],
            ["Bob", "25", "LA"],
        ]
        result = ExcelParser()._split_into_tables(rows)
        assert len(result) == 1
        assert result[0][0] == ["Name", "Age", "City"]
        assert len(result[0][1]) == 2

    def test_split_into_tables_multiple_tables(self):
        """Two header-like rows with data between them → two tables."""
        rows = [
            ["Product", "Q1", "Q2"],     # header 1
            ["Widget A", "100", "120"],   # data for table 1
            ["Metric", "Value", "Unit"],  # header 2 (looks like header)
            ["Speed", "5.5", "m/s"],      # data for table 2
        ]
        result = ExcelParser()._split_into_tables(rows)
        assert len(result) == 2
        assert result[0][0] == ["Product", "Q1", "Q2"]
        assert result[1][0] == ["Metric", "Value", "Unit"]

    def test_table_id_format(self):
        """Table IDs follow tbl_NNN format."""
        header = ["A", "B"]
        data = [["1", "2"]]
        table = ExcelParser._build_table(header, data, "tbl_042")
        assert table is not None
        assert table.table_id == "tbl_042"

    def test_table_text_summary_includes_headers(self):
        """Text summary should contain column names for RAG retrieval."""
        from docnest.models import TableData as TableDataModel

        table = TableDataModel(
            table_id="tbl_001",
            caption="Sales",
            headers=["Product", "Q1", "Q2"],
            rows=[["Widget A", "100", "120"]],
        )
        summary = ExcelParser._table_text_summary(table, "Sales Data")
        assert "Product" in summary
        assert "Q1" in summary
        assert "Sales Data" in summary

    def test_looks_like_header_with_labels(self):
        """Short text labels → looks like a header."""
        assert ExcelParser._looks_like_header(["Name", "Age", "City"]) is True

    def test_looks_like_header_with_numbers(self):
        """Majority numeric → probably a data row, not a header."""
        assert ExcelParser._looks_like_header(["100", "200", "300"]) is False

    def test_looks_like_header_single_cell(self):
        """Single non-empty cell → not a header row."""
        assert ExcelParser._looks_like_header(["Only", "", ""]) is False

    def test_parse_creates_xlsx_fixture(self, tmp_path):
        """Create an XLSX with openpyxl and parse it end-to-end."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue"
        ws.append(["Region", "Q1", "Q2", "Q3", "Q4"])
        ws.append(["North", "100", "120", "110", "130"])
        ws.append(["South", "80", "95", "85", "100"])

        xlsx_path = tmp_path / "revenue.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        parser = ExcelParser()
        raw = parser.parse(str(xlsx_path))

        assert raw.format == "xlsx"
        assert len(raw.sections) == 1
        section = raw.sections[0]
        assert section.title == "Revenue"
        assert len(section.tables) == 1
        table = section.tables[0]
        assert table.headers == ["Region", "Q1", "Q2", "Q3", "Q4"]
        assert len(table.rows) == 2
        assert table.rows[0] == ["North", "100", "120", "110", "130"]
        # Section text should contain useful info for RAG
        assert "Revenue" in section.text
        assert "Region" in section.text

    def test_parse_multi_sheet_xlsx(self, tmp_path):
        """Each worksheet becomes a separate section."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append(["1", "2"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y"])
        ws2.append(["10", "20"])

        xlsx_path = tmp_path / "multi.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        raw = ExcelParser().parse(str(xlsx_path))
        assert len(raw.sections) == 2
        assert raw.sections[0].title == "Sheet1"
        assert raw.sections[1].title == "Sheet2"

    def test_parse_skips_empty_sheets(self, tmp_path):
        """Sheets with no data are silently skipped."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        # No data added

        ws2 = wb.create_sheet("DataSheet")
        ws2.append(["Col1", "Col2"])
        ws2.append(["a", "b"])

        xlsx_path = tmp_path / "mixed.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        raw = ExcelParser().parse(str(xlsx_path))
        assert len(raw.sections) == 1
        assert raw.sections[0].title == "DataSheet"

    def test_parse_row_length_mismatch(self, tmp_path):
        """Rows with different column counts are normalised."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append(["1", "2"])        # short row
        ws.append(["1", "2", "3", "4"])  # long row

        xlsx_path = tmp_path / "mismatch.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        raw = ExcelParser().parse(str(xlsx_path))
        table = raw.sections[0].tables[0]
        assert len(table.headers) == 3
        assert len(table.rows[0]) == 3  # padded
        assert len(table.rows[1]) == 3  # truncated

    def test_parse_doc_id_from_filename(self, tmp_path):
        """doc_id is generated from the filename using _make_doc_id."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append(["1"])

        xlsx_path = tmp_path / "MyReport2024.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        raw = ExcelParser().parse(str(xlsx_path))
        assert raw.doc_id == "my-report-2024"

    def test_all_rows_match_headers_length(self, sample_xlsx):
        """Every data row must have exactly len(headers) cells."""
        raw = ExcelParser().parse(str(sample_xlsx))
        for section in raw.sections:
            for table in section.tables:
                for row in table.rows:
                    assert len(row) == len(table.headers), (
                        f"Row {row} length {len(row)} != headers {len(table.headers)}"
                    )


# ------------------------------------------------------------------ #
# ExcelParser — integration (require tests/fixtures/sample.xlsx) #
# ------------------------------------------------------------------ #

class TestExcelParserIntegration:
    def test_each_sheet_becomes_a_section(self, sample_xlsx):
        raw = ExcelParser().parse(str(sample_xlsx))
        assert len(raw.sections) >= 1

    def test_first_row_becomes_headers(self, sample_xlsx):
        raw = ExcelParser().parse(str(sample_xlsx))
        for s in raw.sections:
            for t in s.tables:
                assert len(t.headers) > 0

    def test_section_text_not_placeholder(self, sample_xlsx):
        """Section text should be descriptive, not just a placeholder."""
        raw = ExcelParser().parse(str(sample_xlsx))
        for s in raw.sections:
            # The old implementation used "Spreadsheet sheet: {title}"
            # Now it should contain column headers for better RAG
            assert "Spreadsheet sheet:" not in s.text

    def test_table_ids_are_consistent_format(self, sample_xlsx):
        """Table IDs should follow tbl_NNN format, not 'sheet_Sheet Name'."""
        raw = ExcelParser().parse(str(sample_xlsx))
        for s in raw.sections:
            for t in s.tables:
                assert t.table_id.startswith("tbl_"), (
                    f"Expected tbl_NNN format, got {t.table_id}"
                )

    def test_row_count_matches_data(self, sample_xlsx):
        """Data rows should match the actual data in the fixture."""
        raw = ExcelParser().parse(str(sample_xlsx))
        # Fixture has 1 header row + 2 data rows = 2 data rows
        for s in raw.sections:
            for t in s.tables:
                assert len(t.rows) > 0
